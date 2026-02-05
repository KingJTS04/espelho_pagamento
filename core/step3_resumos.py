import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment


def gerar_resumos(
    espelhos_xlsx_path: str,
    banco_consolidado_xlsx_path: str,
) -> str:
    """
    Cria as abas RESUMO e RESUMO TOTAL dentro do arquivo Espelhos_Motoristas.xlsx.

    Regras validadas mantidas:
    - RESUMO/RESUMO TOTAL: sempre exibir NOME DO MOTORISTA (se houver prestador, vem de C5 "MOTORISTA: ..."; senão vem de C4 antes do " - documento")
    - Ocultar linhas de grade em TODAS as abas
    - RESUMO TOTAL: adicionar coluna "Valor Bruto" após os clientes e antes do "Desconto"
    - RESUMO TOTAL: adicionar última linha "CUSTO TOTAL" somando cada coluna (clientes, valor bruto, desconto em vermelho, valor líquido)
    - Centralização do nome do motorista e do "CUSTO TOTAL" (aba RESUMO e RESUMO TOTAL)

    Retorna o próprio path do espelhos (arquivo final).
    """
    if not os.path.exists(espelhos_xlsx_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {espelhos_xlsx_path}")
    if not os.path.exists(banco_consolidado_xlsx_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {banco_consolidado_xlsx_path}")

    # =========================
    # ESTILOS
    # =========================
    bold = Font(bold=True)
    bold_red = Font(bold=True, color="FF0000")
    red_font = Font(color="FF0000")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    formato_contabil = 'R$ #,##0.00_);R$ (#,##0.00)'

    # =========================
    # HELPERS
    # =========================
    def norm_text(v):
        return str(v).strip() if v is not None else ""

    def excel_sheet_ref(name: str) -> str:
        return name.replace("'", "''")

    def style_cell(cell, *, font=None, alignment=None, border=None, number_format=None):
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        if number_format is not None:
            cell.number_format = number_format

    def auto_ajuste(ws):
        for column_cells in ws.columns:
            max_length = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 3

    def find_descontos_row(ws):
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                if norm_text(cell.value) == "(-) DESCONTOS":
                    return cell.row
        return None

    def find_valor_bruto_row(ws):
        # Linha onde está o texto do valor da nota, valor fica em F{row}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                texto = str(cell.value)
                if "VALOR TOTAL DOS SERVIÇOS PRESTADOS" in texto:
                    return cell.row
        return None

    def find_mapeamento_header_row(ws):
        for r in range(1, ws.max_row + 1):
            b = norm_text(ws[f"B{r}"].value).upper()
            f = norm_text(ws[f"F{r}"].value).upper()
            if b == "CIDADE" and f == "VALOR TOTAL":
                return r
        return None

    def build_client_ranges_in_mapeamento(ws, clientes_set):
        header_row = find_mapeamento_header_row(ws)
        if not header_row:
            return {}

        ranges = {}
        current_client = None
        start_row = None

        r = header_row + 1
        while r <= ws.max_row:
            s = norm_text(ws[f"B{r}"].value)
            if s:
                sup = s.upper()
                if sup == "TOTAL":
                    break

                if s in clientes_set:
                    if current_client is not None and start_row is not None:
                        end_row = r - 1
                        ranges[current_client] = (start_row, end_row) if end_row >= start_row else None
                    current_client = s
                    start_row = r + 1
            r += 1

        if current_client is not None and start_row is not None:
            end_row = r - 1
            ranges[current_client] = (start_row, end_row) if end_row >= start_row else None

        return ranges

    def nome_limpo(valor):
        """
        Remove sufixo ' - documento' se existir.
        """
        txt = norm_text(valor)
        before, sep, _ = txt.partition(" - ")
        return before.strip() if sep else txt.strip()

    def obter_nome_motorista(ws):
        """
        Regras após mudança do Step2:
        - Se existir prestador: C4 = 'Prestador - doc' e C5 = 'MOTORISTA: Nome'
        - Se NÃO existir prestador: C4 = 'Motorista - doc' e C5 vazio
        Para o RESUMO/RESUMO TOTAL, SEMPRE queremos o NOME DO MOTORISTA.
        """
        c5 = norm_text(ws["C5"].value)
        if c5.upper().startswith("MOTORISTA:"):
            return c5.split(":", 1)[1].strip()
        return nome_limpo(ws["C4"].value)

    # =========================
    # ABRIR ARQUIVOS
    # =========================
    wb_espelhos = load_workbook(espelhos_xlsx_path, data_only=False)
    wb_banco = load_workbook(banco_consolidado_xlsx_path, data_only=True)
    ws_banco = wb_banco.active

    # =========================
    # OCULTAR LINHAS DE GRADE (TODAS AS ABAS)
    # =========================
    for ws in wb_espelhos.worksheets:
        ws.sheet_view.showGridLines = False

    # =========================
    # REMOVER ABAS SE EXISTIREM
    # =========================
    for aba in ("RESUMO", "RESUMO TOTAL"):
        if aba in wb_espelhos.sheetnames:
            del wb_espelhos[aba]

    # =========================
    # CRIAR ABA RESUMO
    # =========================
    ws_resumo = wb_espelhos.create_sheet("RESUMO")
    ws_resumo.sheet_view.showGridLines = False

    ws_resumo["A1"] = "Relação dos Parceiros para Pagamento"
    ws_resumo["A1"].font = bold
    ws_resumo["A2"] = "Centro de Custo:"
    ws_resumo["A3"] = "Período:"
    ws_resumo["E3"] = "Vencimento:"
    ws_resumo["E3"].border = border_all

    linha_inicio = 5
    cabecalhos = ["Nome do motorista", "Valor Bruto", "Desconto", "Valor Líquido", "Status NF"]
    for col, texto in enumerate(cabecalhos, start=1):
        cell = ws_resumo.cell(row=linha_inicio, column=col, value=texto)
        style_cell(
            cell,
            font=bold_red if texto == "Desconto" else bold,
            alignment=center,
            border=border_all
        )

    # =========================
    # PREENCHER MOTORISTAS
    # =========================
    linha_atual = linha_inicio + 1
    linha_primeiro_motorista = linha_atual

    motoristas_list = []
    motorista_to_sheet = {}
    bruto_row_por_motorista = {}
    desconto_row_por_motorista = {}

    for aba in wb_espelhos.sheetnames:
        if aba in ("RESUMO", "RESUMO TOTAL"):
            continue

        ws_m = wb_espelhos[aba]

        motorista = obter_nome_motorista(ws_m)
        if not motorista:
            continue

        bruto_row = find_valor_bruto_row(ws_m)
        desc_row = find_descontos_row(ws_m)

        valor_bruto = ws_m[f"F{bruto_row}"].value if bruto_row else 0
        if valor_bruto is None:
            valor_bruto = 0

        # Nome do motorista centralizado
        cell_nome = ws_resumo.cell(row=linha_atual, column=1, value=motorista)
        style_cell(cell_nome, alignment=center)

        motoristas_list.append(motorista)
        motorista_to_sheet[motorista] = aba
        bruto_row_por_motorista[motorista] = bruto_row
        desconto_row_por_motorista[motorista] = desc_row

        style_cell(
            ws_resumo.cell(row=linha_atual, column=2, value=valor_bruto),
            number_format=formato_contabil,
            alignment=center
        )

        if desc_row:
            c_desc = ws_resumo.cell(row=linha_atual, column=3, value=f"='{aba}'!F{desc_row}")
        else:
            c_desc = ws_resumo.cell(row=linha_atual, column=3, value=0)
        style_cell(c_desc, font=red_font, number_format=formato_contabil, alignment=center)

        style_cell(
            ws_resumo.cell(row=linha_atual, column=4, value=f"=B{linha_atual}-C{linha_atual}"),
            number_format=formato_contabil,
            alignment=center
        )

        style_cell(ws_resumo.cell(row=linha_atual, column=5, value=""), alignment=center)

        for c in range(1, 6):
            ws_resumo.cell(row=linha_atual, column=c).border = border_all

        linha_atual += 1

    # =========================
    # TOTAL RESUMO
    # =========================
    cell_total = ws_resumo.cell(row=linha_atual, column=1, value="CUSTO TOTAL")
    style_cell(cell_total, font=bold, alignment=center, border=border_all)

    style_cell(
        ws_resumo.cell(row=linha_atual, column=2, value=f"=SUM(B{linha_primeiro_motorista}:B{linha_atual-1})"),
        font=bold, number_format=formato_contabil, alignment=center, border=border_all
    )
    style_cell(
        ws_resumo.cell(row=linha_atual, column=3, value=f"=SUM(C{linha_primeiro_motorista}:C{linha_atual-1})"),
        font=bold_red, number_format=formato_contabil, alignment=center, border=border_all
    )
    style_cell(
        ws_resumo.cell(row=linha_atual, column=4, value=f"=SUM(D{linha_primeiro_motorista}:D{linha_atual-1})"),
        font=bold, number_format=formato_contabil, alignment=center, border=border_all
    )

    for c in range(1, 6):
        cell = ws_resumo.cell(row=linha_atual, column=c)
        cell.border = border_all
        if c not in (2, 3, 4):
            cell.font = bold
            cell.alignment = center

    # =========================
    # CRIAR ABA RESUMO TOTAL
    # =========================
    ws_rt = wb_espelhos.create_sheet("RESUMO TOTAL")
    ws_rt.sheet_view.showGridLines = False

    ws_rt["A1"] = "Relação dos Parceiros para Pagamento"
    ws_rt["A1"].font = bold
    ws_rt["A2"] = "Centro de Custo:"
    ws_rt["A3"] = "Período:"

    linha_cab = 5
    linha_rt = 6

    hdr_motor = ws_rt["A5"]
    hdr_motor.value = "Nome do Motorista"
    style_cell(hdr_motor, font=bold, border=border_all, alignment=center)

    for i, motorista in enumerate(motoristas_list):
        cell = ws_rt.cell(row=linha_rt + i, column=1, value=motorista)
        style_cell(cell, border=border_all, alignment=center)

    # Clientes únicos do banco consolidado
    clientes_unicos = []
    col_cliente = None
    for col in range(1, ws_banco.max_column + 1):
        if norm_text(ws_banco.cell(row=1, column=col).value).lower() == "cliente":
            col_cliente = col
            break

    if col_cliente:
        for r in range(2, ws_banco.max_row + 1):
            cliente = ws_banco.cell(row=r, column=col_cliente).value
            if cliente and cliente not in clientes_unicos:
                clientes_unicos.append(cliente)

    clientes_set = set(clientes_unicos)

    col_inicio = 2

    # Cabeçalhos clientes
    for idx, cliente in enumerate(clientes_unicos):
        cell = ws_rt.cell(row=linha_cab, column=col_inicio + idx, value=cliente)
        style_cell(cell, font=bold, alignment=center, border=border_all)

    # NOVA COLUNA: Valor Bruto (após clientes e antes do Desconto)
    col_valor_bruto_rt = col_inicio + len(clientes_unicos)
    style_cell(
        ws_rt.cell(row=linha_cab, column=col_valor_bruto_rt, value="Valor Bruto"),
        font=bold, alignment=center, border=border_all
    )

    # Colunas finais
    col_desconto = col_valor_bruto_rt + 1
    col_liquido = col_desconto + 1
    col_status = col_liquido + 1

    style_cell(ws_rt.cell(row=linha_cab, column=col_desconto, value="Desconto"), font=bold_red, alignment=center, border=border_all)
    style_cell(ws_rt.cell(row=linha_cab, column=col_liquido, value="Valor Líquido"), font=bold, alignment=center, border=border_all)
    style_cell(ws_rt.cell(row=linha_cab, column=col_status, value="Status NF"), font=bold, alignment=center, border=border_all)

    venc = ws_rt.cell(row=3, column=col_status, value="Vencimento:")
    style_cell(venc, font=bold, alignment=center, border=border_all)

    # Pré-cálculos por motorista
    ranges_por_motorista = {}
    for motorista, sheet_name in motorista_to_sheet.items():
        ws_m = wb_espelhos[sheet_name]
        ranges_por_motorista[motorista] = build_client_ranges_in_mapeamento(ws_m, clientes_set)

    bruto_letter = ws_rt.cell(row=linha_cab, column=col_valor_bruto_rt).column_letter
    desc_letter = ws_rt.cell(row=linha_cab, column=col_desconto).column_letter

    # Preencher linhas
    for i, motorista in enumerate(motoristas_list):
        row_out = linha_rt + i
        sheet_name = motorista_to_sheet.get(motorista)
        sheet_ref = excel_sheet_ref(sheet_name) if sheet_name else None

        client_ranges = ranges_por_motorista.get(motorista, {})
        bruto_row = bruto_row_por_motorista.get(motorista)
        desc_row = desconto_row_por_motorista.get(motorista)

        # Clientes
        for j, cliente in enumerate(clientes_unicos):
            out_cell = ws_rt.cell(row=row_out, column=col_inicio + j)
            rng = client_ranges.get(cliente)
            if sheet_ref and rng:
                s, e = rng
                out_cell.value = f"=SUM('{sheet_ref}'!F{s}:F{e})"
                out_cell.number_format = formato_contabil
            else:
                out_cell.value = ""
            out_cell.alignment = center
            out_cell.border = border_all

        # Valor Bruto
        cbruto = ws_rt.cell(row=row_out, column=col_valor_bruto_rt)
        if sheet_ref and bruto_row:
            cbruto.value = f"='{sheet_ref}'!F{bruto_row}"
            cbruto.number_format = formato_contabil
        else:
            cbruto.value = ""
        cbruto.alignment = center
        cbruto.border = border_all

        # Desconto
        cdesc = ws_rt.cell(row=row_out, column=col_desconto)
        if sheet_ref and desc_row:
            cdesc.value = f"='{sheet_ref}'!F{desc_row}"
            cdesc.number_format = formato_contabil
        else:
            cdesc.value = ""
        cdesc.font = red_font
        cdesc.alignment = center
        cdesc.border = border_all

        # Valor Líquido (Valor Bruto - Desconto)
        cliq = ws_rt.cell(row=row_out, column=col_liquido, value=f"={bruto_letter}{row_out}-N({desc_letter}{row_out})")
        style_cell(cliq, number_format=formato_contabil, alignment=center, border=border_all)

        # Status NF (vazio)
        cnf = ws_rt.cell(row=row_out, column=col_status, value="")
        style_cell(cnf, alignment=center, border=border_all)

    # =========================
    # LINHA FINAL CUSTO TOTAL (RESUMO TOTAL)
    # =========================
    row_total_rt = linha_rt + len(motoristas_list)
    first_data_row = linha_rt
    last_data_row = row_total_rt - 1

    cell_total_label = ws_rt.cell(row=row_total_rt, column=1, value="CUSTO TOTAL")
    style_cell(cell_total_label, font=bold, alignment=center, border=border_all)

    # Somar cada coluna numérica (clientes + valor bruto + desconto + valor líquido)
    # clientes
    for c in range(col_inicio, col_inicio + len(clientes_unicos)):
        col_letter = ws_rt.cell(row=linha_cab, column=c).column_letter
        cell = ws_rt.cell(row=row_total_rt, column=c, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
        style_cell(cell, font=bold, alignment=center, border=border_all, number_format=formato_contabil)

    # valor bruto
    col_letter = ws_rt.cell(row=linha_cab, column=col_valor_bruto_rt).column_letter
    cell = ws_rt.cell(row=row_total_rt, column=col_valor_bruto_rt, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
    style_cell(cell, font=bold, alignment=center, border=border_all, number_format=formato_contabil)

    # desconto (vermelho)
    col_letter = ws_rt.cell(row=linha_cab, column=col_desconto).column_letter
    cell = ws_rt.cell(row=row_total_rt, column=col_desconto, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
    style_cell(cell, font=bold_red, alignment=center, border=border_all, number_format=formato_contabil)

    # valor líquido
    col_letter = ws_rt.cell(row=linha_cab, column=col_liquido).column_letter
    cell = ws_rt.cell(row=row_total_rt, column=col_liquido, value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
    style_cell(cell, font=bold, alignment=center, border=border_all, number_format=formato_contabil)

    # status (vazio, mas com borda)
    cell = ws_rt.cell(row=row_total_rt, column=col_status, value="")
    style_cell(cell, font=bold, alignment=center, border=border_all)

    # Garantir borda em toda a linha final (até status)
    for c in range(1, col_status + 1):
        ws_rt.cell(row=row_total_rt, column=c).border = border_all

    # =========================
    # AUTO AJUSTE
    # =========================
    auto_ajuste(ws_resumo)
    auto_ajuste(ws_rt)

    # =========================
    # SALVAR
    # =========================
    wb_espelhos.save(espelhos_xlsx_path)
    return espelhos_xlsx_path