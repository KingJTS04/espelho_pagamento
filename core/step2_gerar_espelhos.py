import os
import re
from io import BytesIO
from typing import Union, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font


FileLike = Union[str, BytesIO, bytes]


def _to_bytes_io(src: FileLike) -> BytesIO:
    """
    Aceita:
      - path (str)
      - bytes
      - BytesIO / file-like

    Retorna BytesIO pronto para leitura (cursor no início).
    """
    if isinstance(src, BytesIO):
        src.seek(0)
        return src
    if isinstance(src, (bytes, bytearray)):
        bio = BytesIO(src)
        bio.seek(0)
        return bio
    with open(src, "rb") as f:
        data = f.read()
    bio = BytesIO(data)
    bio.seek(0)
    return bio


def gerar_espelhos_motoristas(
    banco_consolidado_input: FileLike,
    modelo_input: FileLike,
    *,
    output_dir: Optional[str] = "output/espelhos",
    output_filename: str = "Espelhos_Motoristas.xlsx",
) -> Union[str, bytes]:
    """
    Gera um único XLSX com uma aba por motorista.

    - banco_consolidado_input: path/bytes/BytesIO do banco_consolidado.xlsx
    - modelo_input: path/bytes/BytesIO do modelo.xlsx
    - Se output_dir for None: retorna bytes do XLSX final (ideal para web).
    - Se output_dir for str: salva em disco e retorna o path final.
    """

    # =========================
    # LER BANCO CONSOLIDADO
    # =========================
    banco_io = _to_bytes_io(banco_consolidado_input)
    df = pd.read_excel(banco_io)
    df.columns = df.columns.str.strip().str.lower()

    # =========================
    # IDENTIFICAR COLUNAS
    # =========================
    def achar_coluna(possiveis):
        for c in possiveis:
            if c in df.columns:
                return c
        return None

    col_motorista = achar_coluna(["nome do motorista", "motorista", "nome"])
    col_prestador = achar_coluna(["prestador", "nome do prestador", "prestador de serviço", "prestador de servico"])

    col_conta = achar_coluna(["conta", "conta corrente"])
    col_pix = achar_coluna(["pix", "chave pix"])
    col_data = achar_coluna(["data"])
    col_cidade = achar_coluna(["cidade"])
    col_status = achar_coluna(["status"])
    col_custo = achar_coluna(["custo", "valor", "valor unitario", "valor unitário"])

    # documento (preferir CNPJ, se existir)
    col_cnpj = achar_coluna(["cnpj", "cnpj do favorecido", "cnpj/cpf", "cpf/cnpj"])
    col_cpf = achar_coluna(["cpf", "cpf do favorecido", "cpf/cnpj", "cnpj/cpf"])

    # contrato (vem do motoristas após o merge do step1)
    col_contrato = achar_coluna(["contrato"])

    for obrigatoria in ["cpf", "banco", "agencia", "cliente", "romaneio"]:
        if obrigatoria not in df.columns:
            raise Exception(f"Coluna '{obrigatoria}' não encontrada.")

    if not all([col_motorista, col_conta, col_pix, col_data, col_cidade, col_status, col_custo]):
        raise Exception("Coluna obrigatória não encontrada.")

    # =========================
    # ESTILOS
    # =========================
    align_center = Alignment(horizontal="center", vertical="center")
    align_left_center = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_lr = Border(left=thin, right=thin)
    no_border = Border()

    fill_cliente = PatternFill(fill_type="solid", fgColor="D9D9D9")
    font_bold = Font(bold=True)
    font_bold_red = Font(bold=True, color="FF0000")
    font_red = Font(color="FF0000")

    formato_contabil = "R$ #,##0.00_);R$ (#,##0.00)"

    # =========================
    # HELPERS
    # =========================
    def get_val(v):
        if v is None:
            return ""
        if isinstance(v, float) and pd.isna(v):
            return ""
        return v

    def is_empty(v) -> bool:
        v = get_val(v)
        return str(v).strip() == ""

    def set_info(ws, addr: str, label: str, value):
        """
        Escreve: 'Label: Valor'
        - Se Valor vazio => 'Label: INEXISTENTE' em negrito/vermelho
        - Se Valor preenchido => negrito normal
        """
        if is_empty(value):
            ws[addr] = f"{label}: INEXISTENTE"
            ws[addr].font = font_bold_red
        else:
            ws[addr] = f"{label}: {value}"
            ws[addr].font = font_bold

    def auto_largura_coluna_F(ws, start_row=2, end_row=8):
        max_len = 0
        for r in range(start_row, end_row + 1):
            v = ws[f"F{r}"].value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions["F"].width = max_len + 2

    def aplicar_borda(ws, row, col_ini, col_fim, borda):
        for c in range(col_ini, col_fim + 1):
            ws.cell(row=row, column=c).border = borda

    def aplicar_linha(ws, row, col_ini, col_fim, *, alignment=None, font=None, fill=None, borda=border_all):
        for c in range(col_ini, col_fim + 1):
            cell = ws.cell(row=row, column=c)
            if alignment is not None:
                cell.alignment = alignment
            if font is not None:
                cell.font = font
            if fill is not None:
                cell.fill = fill
            cell.border = borda

    def _sanitize_sheet(name: str) -> str:
        name = re.sub(r"[\\/*?:\[\]]", "", str(name))
        name = re.sub(r"\s+", " ", name).strip()
        return name

    def nome_aba_valido(nome_motorista: str, used: set) -> str:
        """
        Aba = primeiro + segundo nome (se existir), respeitando 31 chars e evitando duplicados.
        """
        base = _sanitize_sheet(nome_motorista)
        parts = [p for p in base.split(" ") if p]
        if len(parts) >= 2:
            base = f"{parts[0]} {parts[1]}"
        elif len(parts) == 1:
            base = parts[0]
        else:
            base = "SEM_NOME"

        base = base[:31].strip()
        candidate = base

        i = 2
        while candidate in used or candidate == "":
            suffix = f" {i}"
            candidate = (base[: max(0, 31 - len(suffix))] + suffix).strip()
            i += 1

        used.add(candidate)
        return candidate

    # =========================
    # ABRIR MODELO (BytesIO compatível)
    # =========================
    modelo_io = _to_bytes_io(modelo_input)
    wb = load_workbook(modelo_io)
    aba_modelo = wb.active
    aba_modelo.title = "MODELO_BASE"

    used_sheet_names = set()

    # =========================
    # GERAR UMA ABA POR MOTORISTA
    # =========================
    for motorista in df[col_motorista].drop_duplicates():
        df_motorista = df[df[col_motorista] == motorista]
        linha_ref = df_motorista.iloc[0]

        ws = wb.copy_worksheet(aba_modelo)
        ws.title = nome_aba_valido(str(motorista), used_sheet_names)

        # =========================
        # PARTE 1 — DADOS FIXOS + PRESTADOR
        # =========================
        prestador = get_val(linha_ref[col_prestador]) if col_prestador else ""
        tem_prestador = str(prestador).strip() != ""

        # Doc: col_cnpj tem prioridade se existir/preenchido, senão cpf.
        # (Aqui assume que o "dono" do doc e dados bancários já vem no banco consolidado conforme sua regra.)
        cnpj_val = get_val(linha_ref[col_cnpj]) if col_cnpj else ""
        cpf_val = get_val(linha_ref[col_cpf]) if col_cpf else get_val(linha_ref["cpf"])
        doc_val = cnpj_val if str(cnpj_val).strip() else cpf_val

        # C4:D4
        if tem_prestador:
            ws["C4"] = f"{prestador} - {doc_val}".strip()
        else:
            ws["C4"] = f"{motorista} - {doc_val}".strip()

        # C5:D5 (só se existir prestador)
        if tem_prestador:
            ws["C5"] = f"MOTORISTA: {motorista}"
            ws["C5"].alignment = align_center
            ws["D5"].alignment = align_center
            # sem bordas
            ws["C5"].border = no_border
            ws["D5"].border = no_border
        else:
            # Garantir vazio quando não tem prestador (não mexe no resto do layout)
            ws["C5"] = ""
            # (não forço borda aqui: mantém o padrão do seu modelo)

        # Contrato
        contrato_val = get_val(linha_ref[col_contrato]) if col_contrato else ""
        if str(contrato_val).strip():
            ws["F2"] = f"Contrato: {contrato_val}"
            ws["F2"].font = font_bold
        else:
            ws["F2"] = "Contrato: INEXISTENTE"
            ws["F2"].font = font_bold_red

        # Dados bancários (mesmas colunas; "dono" já muda conforme prestador estar preenchido no banco consolidado)
        favorecido_nome = prestador if tem_prestador else motorista

        set_info(ws, "F3", "Banco", linha_ref.get("banco"))
        set_info(ws, "F4", "Agência", linha_ref.get("agencia"))
        set_info(ws, "F5", "Conta", linha_ref[col_conta] if col_conta else "")
        set_info(ws, "F6", "Favorecido", favorecido_nome)
        set_info(ws, "F7", "CPF/CNPJ do Favorecido", doc_val)
        set_info(ws, "F8", "PIX", linha_ref[col_pix] if col_pix else "")

        auto_largura_coluna_F(ws, start_row=2, end_row=8)

        # =========================
        # PARTE 2 — TABELA VARIÁVEL
        # =========================
        linha_atual = 11

        for cliente in df_motorista["cliente"].drop_duplicates():
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=6)
            cell = ws.cell(row=linha_atual, column=1)
            cell.value = cliente
            cell.alignment = align_center
            cell.fill = fill_cliente
            cell.font = font_bold
            aplicar_borda(ws, linha_atual, 1, 6, border_all)
            linha_atual += 1

            df_cliente = df_motorista[df_motorista["cliente"] == cliente]
            romaneios = df_cliente["romaneio"].dropna().drop_duplicates()

            for rom in romaneios:
                df_rom = df_cliente[df_cliente["romaneio"] == rom]

                ws.cell(row=linha_atual, column=1, value=rom).font = font_bold
                ws.cell(row=linha_atual, column=2, value=df_rom.shape[0]).font = font_bold
                ws.cell(row=linha_atual, column=4, value=df_rom.iloc[0][col_cidade]).font = font_bold
                ws.cell(row=linha_atual, column=5, value=df_rom.iloc[0][col_data]).font = font_bold
                ws.cell(row=linha_atual, column=6, value=df_rom.iloc[0][col_status]).font = font_bold

                aplicar_linha(ws, linha_atual, 1, 6, alignment=align_center, borda=border_all)
                linha_atual += 1

        # =========================
        # PARTE 3 — MAPEAMENTO POR CIDADE
        # =========================
        linha_atual += 1

        ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=3)
        ws[f"B{linha_atual}"] = "CIDADE"
        ws[f"D{linha_atual}"] = "QUANTIDADE"
        ws[f"E{linha_atual}"] = "VALOR UNITÁRIO"
        ws[f"F{linha_atual}"] = "VALOR TOTAL"

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{linha_atual}"].font = font_bold
            ws[f"{col}{linha_atual}"].alignment = align_center
            ws[f"{col}{linha_atual}"].border = border_all

        linha_atual += 1

        soma_geral_qtd = 0
        soma_geral_valor = 0

        for cliente in df_motorista["cliente"].drop_duplicates():
            ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=6)
            ws[f"B{linha_atual}"] = cliente
            ws[f"B{linha_atual}"].font = font_bold
            ws[f"B{linha_atual}"].alignment = align_center
            ws[f"B{linha_atual}"].fill = fill_cliente

            for col in ["B", "C", "D", "E", "F"]:
                ws[f"{col}{linha_atual}"].border = border_all

            linha_atual += 1

            df_cliente = df_motorista[df_motorista["cliente"] == cliente]
            cidades = df_cliente[col_cidade].dropna().unique()

            for cidade in cidades:
                quantidade = df_cliente[df_cliente[col_cidade] == cidade].shape[0]
                valor_unitario = df_cliente[df_cliente[col_cidade] == cidade].iloc[0][col_custo]
                valor_total = quantidade * valor_unitario

                soma_geral_qtd += quantidade
                soma_geral_valor += valor_total

                ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=3)
                ws[f"B{linha_atual}"] = cidade
                ws[f"D{linha_atual}"] = quantidade
                ws[f"E{linha_atual}"] = valor_unitario
                ws[f"F{linha_atual}"] = valor_total

                ws[f"E{linha_atual}"].number_format = formato_contabil
                ws[f"F{linha_atual}"].number_format = formato_contabil

                for col in ["B", "C", "D", "E", "F"]:
                    ws[f"{col}{linha_atual}"].alignment = align_center
                    ws[f"{col}{linha_atual}"].border = border_all

                linha_atual += 1

        ws.merge_cells(start_row=linha_atual, start_column=2, end_row=linha_atual, end_column=3)
        ws[f"B{linha_atual}"] = "TOTAL"
        ws[f"D{linha_atual}"] = soma_geral_qtd
        ws[f"E{linha_atual}"] = "-"
        ws[f"F{linha_atual}"] = soma_geral_valor
        ws[f"F{linha_atual}"].number_format = formato_contabil

        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{linha_atual}"].font = font_bold
            ws[f"{col}{linha_atual}"].alignment = align_center
            ws[f"{col}{linha_atual}"].border = border_all

        # =========================
        # VALOR TOTAL DA NOTA
        # =========================
        linha_atual += 2
        linha_valor_nota = linha_atual

        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=5)
        ws[f"A{linha_atual}"] = "VALOR TOTAL DOS SERVIÇOS PRESTADOS NO PERÍODO (valor da nota fiscal)"
        ws[f"A{linha_atual}"].font = font_bold
        ws[f"A{linha_atual}"].fill = fill_cliente
        ws[f"A{linha_atual}"].alignment = align_left_center

        ws[f"F{linha_atual}"] = soma_geral_valor
        ws[f"F{linha_atual}"].font = font_bold
        ws[f"F{linha_atual}"].fill = fill_cliente
        ws[f"F{linha_atual}"].alignment = align_center
        ws[f"F{linha_atual}"].number_format = formato_contabil

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{linha_atual}"].border = border_all

        # =========================
        # PARTE 4 — DESCONTOS E VALOR LÍQUIDO
        # =========================
        linha_atual += 2
        linha_descontos = linha_atual

        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=5)
        ws[f"A{linha_atual}"] = "(-) DESCONTOS"
        ws[f"A{linha_atual}"].font = font_bold
        ws[f"A{linha_atual}"].alignment = align_left_center

        ws[f"F{linha_atual}"] = f"=SUM(F{linha_atual+1}:F{linha_atual+5})"
        ws[f"F{linha_atual}"].number_format = formato_contabil
        ws[f"F{linha_atual}"].font = font_bold_red
        ws[f"F{linha_atual}"].alignment = align_center

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{linha_atual}"].border = border_all

        for _ in range(5):
            linha_atual += 1
            ws[f"A{linha_atual}"].border = border_lr
            ws[f"F{linha_atual}"].border = border_lr

            ws[f"F{linha_atual}"].alignment = align_center
            ws[f"F{linha_atual}"].font = font_red
            ws[f"F{linha_atual}"].number_format = formato_contabil

        linha_atual += 1

        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=5)
        ws[f"A{linha_atual}"] = "VALOR LÍQUIDO A PAGAR AO PRESTADOR DE SERVIÇO"
        ws[f"A{linha_atual}"].font = font_bold
        ws[f"A{linha_atual}"].alignment = align_left_center

        ws[f"F{linha_atual}"] = f"=F{linha_valor_nota}-F{linha_descontos}"
        ws[f"F{linha_atual}"].number_format = formato_contabil
        ws[f"F{linha_atual}"].font = font_bold
        ws[f"F{linha_atual}"].alignment = align_center

        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{linha_atual}"].border = border_all
            ws[f"{col}{linha_atual}"].fill = fill_cliente
            ws[f"{col}{linha_atual}"].font = font_bold

    # =========================
    # REMOVER ABA MODELO
    # =========================
    del wb["MODELO_BASE"]

    # =========================
    # SALVAR (disco ou bytes)
    # =========================
    if output_dir is None:
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out.getvalue()

    os.makedirs(output_dir, exist_ok=True)
    final_path = os.path.join(output_dir, output_filename)
    wb.save(final_path)
    return final_path


# =========================
# USO OFFLINE
# =========================
if __name__ == "__main__":
    CAMINHO_BANCO = "output/banco_consolidado.xlsx"
    CAMINHO_MODELO = "modelo/modelo.xlsx"
    PASTA_ESPELHOS = "output/espelhos"
    ARQUIVO_FINAL = "Espelhos_Motoristas.xlsx"

    saida = gerar_espelhos_motoristas(
        CAMINHO_BANCO,
        CAMINHO_MODELO,
        output_dir=PASTA_ESPELHOS,
        output_filename=ARQUIVO_FINAL,
    )
    print("✅ Espelhos gerados:", saida)